from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from scripts.export_records import to_xlsx_bytes, ensure_records, all_columns


def test_to_xlsx_bytes_roundtrip():
    data = [
        {"a": 1, "b": 2},
        {"a": 3, "c": 4},
    ]
    records = ensure_records(data)
    cols = all_columns(records)

    blob = to_xlsx_bytes(records, cols)
    assert isinstance(blob, (bytes, bytearray)) and len(blob) > 0

    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    # Header row
    headers = [cell.value for cell in ws[1]]
    assert headers == cols
    # First data row
    row2 = [cell.value for cell in ws[2]]
    assert row2[0] == 1 and row2[1] == 2
    # Second data row, third column value present for 'c'
    row3 = [cell.value for cell in ws[3]]
    assert row3[0] == 3 and row3[2] == 4

