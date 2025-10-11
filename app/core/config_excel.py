from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from app.core.config_schema import AppModelConfig, PricingConfig


class ExcelConfigError(RuntimeError):
    pass


_BOOL_TRUE = {"y", "yes", "true", "1", "t"}
_BOOL_FALSE = {"n", "no", "false", "0", "f"}


def _to_snake(value: str) -> str:
    return value.strip().lower().replace(" ", "_") if value else ""


def _is_truthy(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    stringified = str(value).strip().lower()
    if stringified in _BOOL_TRUE:
        return True
    if stringified in _BOOL_FALSE:
        return False
    return None


def _as_bool(value: Any, default: bool = False) -> bool:
    flag = _is_truthy(value)
    return default if flag is None else flag


def _as_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ExcelConfigError(f"Could not parse float from value '{value}'")


def _as_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise ExcelConfigError(f"Could not parse int from value '{value}'")


def _as_csv_ints(value: Any) -> List[int]:
    if value is None or value == "":
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    items = []
    for token in str(value).replace(";", ",").split(","):
        token = token.strip()
        if not token:
            continue
        try:
            items.append(int(token))
        except ValueError:
            raise ExcelConfigError(f"Could not parse integer from '{token}' in list '{value}'")
    return items


def _as_csv_strings(value: Any) -> List[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    tokens = [token.strip() for token in str(value).replace(";", ",").split(",")]
    return [token for token in tokens if token]


@dataclass
class _Row:
    cells: Dict[str, Any]
    raw_index: int

    def get(self, key: str, default: Any = None) -> Any:
        return self.cells.get(key, default)


def _rows_from_sheet(path: Path, sheet_name: str) -> List[_Row]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ExcelConfigError(f"Missing sheet '{sheet_name}' in {path}")
    sheet = wb[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_to_snake(str(cell) if cell is not None else "") for cell in header_row]
    rows: List[_Row] = []
    for idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in values):
            continue
        mapped = {headers[col]: values[col] for col in range(min(len(headers), len(values)))}
        rows.append(_Row(mapped, raw_index=idx))
    wb.close()
    return rows


def _row_applies(row_profile: Optional[str], profile: str) -> bool:
    if row_profile is None or str(row_profile).strip() == "":
        return True
    normalized = str(row_profile).strip().lower()
    if normalized in {"all", "*"}:
        return True
    return normalized == profile.lower()


def parse_excel_config(path: Path, profile: str) -> AppModelConfig:
    profile_key = profile.lower()

    profile_rows = _rows_from_sheet(path, "profiles")
    profile_map: Dict[str, _Row] = {}
    for row in profile_rows:
        row_profile = str(row.get("profile", "")).strip().lower()
        if not row_profile:
            continue
        profile_map[row_profile] = row
    if profile_key not in profile_map:
        raise ExcelConfigError(f"Profile '{profile}' not defined in profiles sheet")

    prof = profile_map[profile_key]
    defaults_provider = prof.get("default_provider")
    defaults_model = prof.get("default_model")
    if not defaults_provider or not defaults_model:
        raise ExcelConfigError(f"Profile '{profile}' must set default_provider and default_model")

    policies = {
        "max_concurrent_requests": _as_int(prof.get("max_concurrent_requests")) or 4,
        "allow_frontend_model_selection": _as_bool(prof.get("allow_frontend_model_selection"), False),
        "allow_frontend_temperature": _as_bool(prof.get("allow_frontend_temperature"), False),
        "redaction": {
            "enabled": _as_bool(prof.get("redaction_enabled"), False),
            "fields": _as_csv_strings(prof.get("redaction_fields")),
        },
    }

    provider_rows = _rows_from_sheet(path, "providers")
    providers_by_id: Dict[str, Dict[str, Any]] = {}
    for row in provider_rows:
        row_profile = row.get("profile")
        if not _row_applies(row_profile, profile):
            continue
        provider_id = row.get("provider_id") or row.get("id")
        if not provider_id:
            raise ExcelConfigError(f"Provider row {row.raw_index} missing provider_id")
        provider_id = str(provider_id).strip()
        entry = {
            "id": provider_id,
            "label": row.get("label") or provider_id,
            "type": (row.get("type") or "openai_compatible").strip(),
            "base_url": (row.get("base_url") or "").strip(),
            "enabled": _as_bool(row.get("enabled"), True),
            "auth": {
                "mode": (row.get("auth_mode") or "bearer_token").strip(),
                "token_source": (row.get("token_source") or "env").strip() or None,
                "token_ref": (row.get("token_ref") or "").strip() or None,
                "header_name": (row.get("header_name") or "").strip() or None,
                "query_param_name": (row.get("query_param_name") or "").strip() or None,
            },
            "timeouts": {
                "connect_s": _as_float(row.get("connect_timeout_s")) or 5.0,
                "read_s": _as_float(row.get("read_timeout_s")) or 60.0,
                "total_s": _as_float(row.get("total_timeout_s")) or 120.0,
            },
            "retry": {
                "max_retries": _as_int(row.get("max_retries")) or 2,
                "backoff_s": _as_float(row.get("backoff_s")) or 0.5,
                "retry_on": _as_csv_ints(row.get("retry_http_codes")),
            },
            "models": [],
        }
        if not entry["base_url"]:
            raise ExcelConfigError(f"Provider '{provider_id}' missing base_url")
        providers_by_id[provider_id] = entry

    if defaults_provider not in providers_by_id:
        raise ExcelConfigError(
            f"Default provider '{defaults_provider}' not present for profile '{profile}'"
        )

    model_rows = _rows_from_sheet(path, "models")
    for row in model_rows:
        if not _row_applies(row.get("profile"), profile):
            continue
        provider_id = row.get("provider_id")
        model_id = row.get("model_id") or row.get("id")
        if not provider_id or not model_id:
            raise ExcelConfigError(f"Model row {row.raw_index} missing provider_id or model_id")
        provider_id = str(provider_id).strip()
        model_id = str(model_id).strip()
        provider = providers_by_id.get(provider_id)
        if provider is None:
            raise ExcelConfigError(
                f"Model '{model_id}' references unknown provider '{provider_id}' for profile '{profile}'"
            )

        price_input_per_million = _as_float(row.get("price_input_per_million"))
        price_output_per_million = _as_float(row.get("price_output_per_million"))
        price_cache_read_per_million = _as_float(row.get("price_cache_read_per_million"))
        price_cache_write_per_million = _as_float(row.get("price_cache_write_per_million"))

        def _per_1k(value: Optional[float]) -> Optional[float]:
            if value is None:
                return None
            return value / 1000.0

        pricing = {
            "currency": (row.get("currency") or "USD").strip(),
            "input_per_1k": _per_1k(price_input_per_million) or 0.0,
            "output_per_1k": _per_1k(price_output_per_million) or 0.0,
            "cache_read_per_1k": _per_1k(price_cache_read_per_million),
            "cache_write_per_1k": _per_1k(price_cache_write_per_million),
            "input_per_million": price_input_per_million,
            "output_per_million": price_output_per_million,
            "cache_read_per_million": price_cache_read_per_million,
            "cache_write_per_million": price_cache_write_per_million,
        }

        capabilities = {
            "json_mode": _as_bool(row.get("cap_json_mode"), False),
            "structured_output": _as_bool(row.get("cap_structured_output"), False),
            "tools": _as_bool(row.get("cap_tools"), False),
            "parallel_tool_calls": _as_bool(row.get("cap_parallel_tools"), False),
            "streaming": _as_bool(row.get("cap_streaming"), False),
            "vision": _as_bool(row.get("cap_vision"), False),
            "audio_in": _as_bool(row.get("cap_audio_in"), False),
            "audio_out": _as_bool(row.get("cap_audio_out"), False),
            "embeddings": _as_bool(row.get("cap_embeddings"), False),
        }

        compatibility = {
            "image_part_key": (row.get("compat_image_part_key") or "").strip() or None,
            "tool_schema_style": (row.get("compat_tool_schema_style") or "").strip() or None,
            "extra_headers": {},
            "max_tokens_param": (row.get("compat_max_tokens_param") or "").strip() or None,
            "allow_input_image_fallback": _as_bool(row.get("compat_allow_input_image"), True),
        }

        show_in_ui = _as_bool(row.get("show_in_ui"), True)
        allow_temp_override = _as_bool(row.get("allow_frontend_override_temperature"), True)
        allow_reasoning_override = _as_bool(row.get("allow_frontend_override_reasoning"), True)

        reasoning = {
            "provider": (row.get("reasoning_provider") or "").strip() or None,
            "effort_default": (row.get("reasoning_effort_default") or "").strip() or None,
            "include_thoughts_default": _as_bool(row.get("include_thoughts_default"), False),
            "allow_override": allow_reasoning_override,
        }

        model_entry = {
            "id": model_id,
            "label": row.get("label") or model_id,
            "route": (row.get("route") or "chat_completions").strip(),
            "context_window": _as_int(row.get("context_window")) or 8192,
            "max_output_tokens": _as_int(row.get("max_output_tokens")),
            "max_temperature": _as_float(row.get("max_temperature")) or 1.0,
            "default_temperature": _as_float(row.get("default_temperature")) or 0.7,
            "default_top_p": _as_float(row.get("default_top_p")),
            "force_json_mode": _as_bool(row.get("force_json_mode"), False),
            "prefer_tools": _as_bool(row.get("prefer_tools"), False),
            "pricing": pricing,
            "capabilities": capabilities,
            "compatibility": compatibility,
            "reasoning": reasoning,
            "show_in_ui": show_in_ui,
            "allow_frontend_override_temperature": allow_temp_override,
            "allow_frontend_override_reasoning": allow_reasoning_override,
        }

        providers_models: List[Dict[str, Any]] = provider.setdefault("models", [])
        providers_models.append(model_entry)

    providers_list = list(providers_by_id.values())
    config_dict = {
        "version": 1,
        "profile": profile,
        "defaults": {
            "provider": defaults_provider,
            "model": defaults_model,
        },
        "policies": policies,
        "providers": providers_list,
    }

    return AppModelConfig.model_validate(config_dict)
